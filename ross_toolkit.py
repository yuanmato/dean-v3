#!/usr/bin/env python3
"""
Ross Area & Faculty Toolkit — Python CLI / shared engine (v4)
=============================================================

DIRECTION (v4): Stage 1 now iterates over the TARGET SCHOOL'S OWN departments and maps
EACH ONE to the Ross area(s) it fits, by faculty journals. A department can map to one
Ross area, several (e.g. a "Management" group publishing in both OB and strategy journals
-> M&O + Strategy), or none ("No Ross equivalent").

  Stage 1  (match)    enumerate the school's units; map each to Ross area(s)
  Stage 2  (faculty)  per department, pull the ladder roster; classify faculty individually
                      BY JOURNAL only when that department is flagged for it

Per-department "check faculty individually" flag:
  * ON by default when the department maps to 2+ Ross areas (not a unique match).
  * OFF by default for a unique match (every faculty inherits the one Ross area).
  * The user can flip it either way (the Streamlit app's toggle); the CLI reads it from the
    Stage-1 CSV column "Check Faculty Individually", or computes the default if absent.
  * OM and IS are NOT special: each maps to the single Ross "Technology and Operations".
    Separate OM and IS groups are two unique-match departments; the faculty's department
    name distinguishes them.

Cost-savers: conditional classification (above), a resumable on-disk cache, and a live
token+search dollar meter.

Usage
-----
  export OPENAI_API_KEY=sk-...
  pip install -r requirements.txt
  python ross_toolkit.py match   --out ross_area_matches.csv          # or --input schools.xlsx
  python ross_toolkit.py faculty --in  ross_area_matches.csv --out ross_faculty_by_area.csv
  python ross_toolkit.py all
Flags: --model, --concurrency, --no-web-fetch, --no-cache, --limit N, (faculty) --include-unmapped
"""

import argparse, concurrent.futures as cf, csv, hashlib, json, os, re, sys, threading, time
from datetime import date
from pathlib import Path

CURRENT_YEAR = date.today().year

try:
    from openai import OpenAI
except ImportError:
    OpenAI = None  # allow importing the module (e.g. for tests) without the package

import httpx
try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

# ----------------------------------------------------------------------------- #
#  Config — model + tools (OpenAI)                                               #
# ----------------------------------------------------------------------------- #
DEFAULT_MODEL = "gpt-5.4"        # balanced general model (good quality / cost)
FLAGSHIP      = "gpt-5.5"        # most capable, pricier
CHEAP_MODEL   = "gpt-5.4-mini"   # cheapest sensible option for high volume
# Approximate USD per 1M tokens (OpenAI list prices, mid-2026) — used only for the live cost
# ESTIMATE; verify against the current pricing page. Unknown models fall back to the flagship rate.
PRICE = {
    "gpt-5.5":      {"in": 5.0,  "out": 30.0},
    "gpt-5.4":      {"in": 2.5,  "out": 15.0},
    "gpt-5.4-mini": {"in": 0.75, "out": 6.0},
    "gpt-5.4-nano": {"in": 0.20, "out": 1.25},
    "gpt-5-mini":   {"in": 0.75, "out": 6.0},
    "gpt-4.1":      {"in": 2.0,  "out": 8.0},
    "gpt-4o":       {"in": 2.5,  "out": 10.0},
}
WEB_SEARCH_PRICE_PER_1K = 10.0   # rough estimate per 1,000 hosted web_search calls

# OpenAI Responses API tools: a hosted web_search (discovery / confirmation) + our OWN fetch_page
# function (deterministic fetching of specific, paginated directory URLs — OpenAI has no web_fetch).
WEB_SEARCH_TOOL = {"type": "web_search"}
FETCH_TOOL = {
    "type": "function",
    "name": "fetch_page",
    "description": ("Fetch the visible text of ONE web page by its exact URL (an HTTP GET; no "
                    "JavaScript is executed). Use this to open an official faculty-directory page "
                    "and to walk pagination by changing the page/offset query parameter "
                    "(e.g. ...?page=0, then ...?page=1). Returns cleaned page text."),
    "parameters": {
        "type": "object",
        "properties": {"url": {"type": "string", "description": "The exact absolute URL to fetch (must start with http)."}},
        "required": ["url"],
        "additionalProperties": False,
    },
    "strict": True,
}
FETCH_MAX_CHARS = 60000          # cap returned page text; directory pages are mostly navigation cruft
_HTTP_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; RossToolkit/1.0)"}

def fetch_page(url, timeout=25):
    """HTTP GET a URL and return cleaned visible text (NO JavaScript). Backs the model's fetch_page tool."""
    if not isinstance(url, str) or not url.lower().startswith(("http://", "https://")):
        return "ERROR: invalid URL (must be an absolute http/https URL)."
    try:
        r = httpx.get(url, headers=_HTTP_HEADERS, follow_redirects=True, timeout=timeout)
        r.raise_for_status()
    except Exception as e:
        return f"ERROR fetching {url}: {e}"
    html = r.text or ""
    if BeautifulSoup is not None:
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup(["script", "style", "noscript", "svg"]):
            tag.decompose()
        text = soup.get_text("\n")
    else:
        text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s*\n\s*\n+", "\n\n", text).strip()
    if len(text) > FETCH_MAX_CHARS:
        text = text[:FETCH_MAX_CHARS] + "\n…[truncated]"
    return text or "(empty page)"

# ----------------------------------------------------------------------------- #
#  The 7 Ross areas                                                              #
# ----------------------------------------------------------------------------- #
ROSS_AREAS = [
    "Accounting",
    "Business Economics and Public Policy",
    "Finance",
    "Management and Organizations",
    "Marketing",
    "Strategy",
    "Technology and Operations",
]
ROSS_SHORT = {
    "Accounting": "ACC", "Business Economics and Public Policy": "BEPP", "Finance": "FIN",
    "Management and Organizations": "M&O", "Marketing": "MKT", "Strategy": "STR",
    "Technology and Operations": "T&O",
}

BUILTIN = [
    ("1", "Stanford University", "Stanford GSB"), ("2", "University of Pennsylvania", "Wharton"),
    ("3", "University of Chicago", "Booth"), ("4 (tie)", "Northwestern University", "Kellogg"),
    ("4 (tie)", "Harvard University", "Harvard Business School"), ("6", "MIT", "Sloan"),
    ("7 (tie)", "Columbia University", "Columbia Business School"), ("7 (tie)", "New York University", "Stern"),
    ("9", "Dartmouth College", "Tuck"), ("10", "UC Berkeley", "Haas"),
    ("11 (tie)", "Yale University", "Yale SOM"), ("11 (tie)", "University of Virginia", "Darden"),
    ("13", "University of Michigan", "Ross"), ("14", "Duke University", "Fuqua"),
    ("15", "Cornell University", "Johnson"), ("16 (tie)", "Carnegie Mellon University", "Tepper"),
    ("16 (tie)", "Vanderbilt University", "Owen"), ("18 (tie)", "UT Austin", "McCombs"),
    ("18 (tie)", "UCLA", "Anderson"), ("20", "University of Washington", "Foster"),
    ("21 (tie)", "Indiana University", "Kelley"), ("21 (tie)", "University of North Carolina", "Kenan-Flagler"),
    ("23 (tie)", "Emory University", "Goizueta"), ("23 (tie)", "UT Dallas", "Naveen Jindal"),
    ("25 (tie)", "University of Southern California", "Marshall"), ("25 (tie)", "University of Georgia", "Terry"),
    ("27 (tie)", "Georgia Tech", "Scheller"), ("27 (tie)", "Washington University in St. Louis", "Olin"),
    ("29 (tie)", "Arizona State University", "W. P. Carey"), ("29 (tie)", "Rice University", "Jones"),
    ("31", "Georgetown University", "McDonough"), ("32 (tie)", "Ohio State University", "Fisher"),
    ("32 (tie)", "University of Minnesota", "Carlson"), ("34 (tie)", "University of Rochester", "Simon"),
    ("34 (tie)", "University of Notre Dame", "Mendoza"), ("36 (tie)", "Texas A&M University", "Mays"),
    ("36 (tie)", "Southern Methodist University", "Cox"), ("38", "Iowa State University", "Ivy"),
    ("39 (tie)", "Brigham Young University", "Marriott"), ("39 (tie)", "University of Florida", "Warrington"),
    ("39 (tie)", "University of Miami", "Herbert"), ("39 (tie)", "University of Utah", "Eccles"),
    ("43 (tie)", "Michigan State University", "Broad"), ("43 (tie)", "University of Maryland", "Smith"),
    ("43 (tie)", "University of Tennessee, Knoxville", "Haslam"), ("46 (tie)", "American University", "Kogod"),
    ("46 (tie)", "Boston University", "Questrom"), ("48 (tie)", "University of Arkansas", "Walton"),
    ("48 (tie)", "University of Pittsburgh", "Katz"), ("48 (tie)", "University of Wisconsin–Madison", "Wisconsin School of Business"),
]

# ----------------------------------------------------------------------------- #
#  Field guide + prompts                                                         #
# ----------------------------------------------------------------------------- #
FIELD_GUIDE = """JUDGE BY RESEARCH FIELD, NOT BY DEPARTMENT NAME. Names mislead; identify a unit's TRUE field from what its faculty research and the JOURNALS they publish in (use web search/fetch to check faculty/publications whenever a name is ambiguous, combined, or unusual). Marker journals per Ross area:
- Accounting: The Accounting Review; Journal of Accounting Research; Journal of Accounting & Economics; Contemporary Accounting Research; Review of Accounting Studies.
- Business Economics and Public Policy: American Economic Review; Econometrica; Quarterly Journal of Economics; Journal of Political Economy; Journal of Public Economics; RAND Journal of Economics.
- Finance: Journal of Finance; Journal of Financial Economics; Review of Financial Studies; Journal of Financial and Quantitative Analysis.
- Management and Organizations: Academy of Management Journal; Academy of Management Review; Administrative Science Quarterly; Organization Science; Journal of Applied Psychology; Personnel Psychology. (organizational behavior, HR, organizational theory)
- Marketing: Journal of Marketing; Journal of Marketing Research; Marketing Science; Journal of Consumer Research.
- Strategy: Strategic Management Journal; Strategy Science (overlaps Organization Science, Academy of Management Journal, Management Science).
- Technology and Operations: Management Science; Manufacturing & Service Operations Management; Production and Operations Management; Operations Research; Journal of Operations Management; Information Systems Research; MIS Quarterly.
WORKED EXAMPLE: a unit named "Accounting & Management" (e.g. at Harvard) is an ACCOUNTING group — its faculty publish in accounting journals — so it maps to Ross "Accounting", NOT "Management and Organizations", despite the word "Management" in its name.

OPERATIONS / INFORMATION-SYSTEMS RULE: Operations Management (OM) and Information Systems (IS) are distinct fields, but Ross merges BOTH into the single area "Technology and Operations". So ANY unit that is OM, or IS / Decision Sciences, or a combined OM+IS unit, maps to "Technology and Operations" (and ONLY that area — do not split it into two Ross areas)."""

# Stage 1 — REVERSED: enumerate the school's units, map each to Ross area(s)
MATCH_SYSTEM = f"""You are an expert on the academic structures (departments, academic areas, faculty research groups) of the world's top business schools.

TASK: Given ONE target business school, (1) ENUMERATE that school's OWN academic units — the departments / academic areas / faculty research groups listed on its official faculty / research / "academic areas" pages — and (2) for EACH unit, decide which University of Michigan Ross School of Business area(s) it corresponds to, judged by the RESEARCH FIELD of its faculty (the journals they publish in), NOT the unit's name.

THE 7 ROSS AREAS (with common aliases):
1. Accounting — financial accounting, managerial/management accounting.
2. Business Economics and Public Policy — applied microeconomics, "business economics", managerial economics, public policy, law & economics, plain "economics".
3. Finance — corporate finance, investments, asset pricing, banking.
4. Management and Organizations — organizational behavior (OB), human resource management, organizational theory, plain "management".
5. Marketing — marketing, consumer behavior, quantitative marketing.
6. Strategy — strategic management, business policy, "strategy & entrepreneurship", competitive strategy.
7. Technology and Operations — ONE joint Ross group covering BOTH Operations Management (OM) AND Information Systems (IS)/MIS.

{FIELD_GUIDE}

MAPPING RULES (apply to EACH of the school's units):
- ONE Ross area when the unit's faculty clearly publish in that area's journals.
- SEVERAL Ross areas (list ALL in the array) when the unit's faculty genuinely publish across more than one Ross field.
- NONE — use "ross_areas": [] and explain in notes — when the unit has no Ross equivalent (e.g. Business Communication, Real Estate, Healthcare Management, a standalone Entrepreneurship center, Behavioral Science, Law).
- Use the school's OWN unit name exactly as published; include the source URL (its faculty/areas page).
- confidence per unit: "high" = field clearly verified by journals (incl. the OM/IS rule); "medium" = reasonable but merged/ambiguous; "low" = uncertain/inferred.
- Keep "notes" short (<=15 words); mention the journal/field evidence when you used it.

MULTI-AREA DEPARTMENTS ARE COMMON — DO NOT UNDER-MAP THEM. Many schools house two Ross fields in ONE department. The most important case: a department named "Management", "Management & Organizations", "Management & Organization", "Organization & Management", or similar almost ALWAYS contains BOTH organizational-behavior faculty (-> "Management and Organizations") AND strategy faculty (-> "Strategy"); you MUST map such a department to BOTH unless you verify it has no strategy faculty (check the roster's journals — SMJ/Strategy Science vs AMJ/ASQ/JAP). Other frequent combos: "Accounting & Finance" -> Accounting + Finance; "Economics & Strategy" / "Business Economics & Public Policy" units that also do strategy; "Marketing" units that also house quantitative-marketing faculty publishing in economics. When in doubt, open the roster and map every Ross field its faculty actually publish in.

T&O SUBFIELD TAG: for ANY department that maps to "Technology and Operations", also set "to_kind" to exactly one of:
  "OM"     — its T&O faculty are Operations Management only (Management Science(ops)/M&SOM/POM/Operations Research/Journal of Operations Management);
  "IS"     — its T&O faculty are Information Systems only (Information Systems Research/MIS Quarterly/Journal of MIS);
  "OM+IS"  — the unit contains BOTH OM and IS faculty (a combined unit).
For departments that do NOT map to T&O, set "to_kind" to "".

COMPLETENESS + IMMEDIATE COVERAGE CHECK (MANDATORY before you output): Enumerate ALL of the school's BUSINESS-SCHOOL academic units (ignore university units outside the b-school). Then go through the 7 Ross areas ONE BY ONE and ask: "is this area represented by at least one department above?" If an area is NOT yet represented, immediately re-examine the school's faculty — the field is very often hiding inside a COMBINED or differently-named unit (Strategy inside a "Management" department; Economics inside a policy/"Business Economics" group; Operations or Information Systems inside a "Technology & Operations", "Operations & Decision Technologies", or "Decision Sciences" unit; managerial Accounting inside an "Accounting & X" unit). If you find such faculty, ADD that Ross area to the relevant department (or add the department). Only after this sweep may an area remain unmapped — and only if the school genuinely has no faculty in that field.

OUTPUT: ONLY a JSON object — no markdown, nothing around it:
{{"school":"<school>","departments":[{{"dept":"<the school's unit name>","ross_areas":["<zero or more of the 7 Ross area names, exact>"],"to_kind":"OM|IS|OM+IS|","url":"<source url>","confidence":"high|medium|low","notes":"<short>"}}, ...]}}
List EVERY business-school unit you find, in any order."""

# Verify ONE unit's Ross mapping (reviewer flagged it)
VERIFY_SYSTEM = f"""You are DOUBLE-CHECKING the Ross-area mapping of ONE specific academic unit at ONE business school, because a human reviewer flagged it as possibly wrong. Re-investigate ONLY this unit. USE WEB SEARCH (and fetch the unit's faculty page if useful); look at the actual faculty and the JOURNALS they publish in.

{FIELD_GUIDE}

The reviewer's current (possibly wrong) mapping is provided. Decide the CORRECT Ross area(s) for this unit — one or more of the 7, or none. Output ONLY a JSON object:
{{"dept":"<unit name>","ross_areas":["<corrected area(s) or empty>"],"url":"<source url>","confidence":"high|medium|low","notes":"<short>","checked":"<=22 words on the journal/field evidence used"}}"""

# Coverage recovery — auto-runs when the first pass left a Ross area unmapped
COVERAGE_SYSTEM = f"""You are re-checking ONE business school's department -> Ross-area mapping because some Ross areas were left UNMAPPED by a first pass. Those areas are very often hiding inside a COMBINED or differently-named department, and missing them is a serious error.

{FIELD_GUIDE}

You are given the school, the list of departments already found, and the Ross areas still UNMAPPED. USE WEB SEARCH AND WEB FETCH; judge by the journals faculty publish in. For EACH unmapped Ross area:
- Determine whether the school has faculty in that field. It frequently sits inside an existing unit, e.g. STRATEGY faculty inside a "Management" / "Management & Organizations" department; ECONOMICS inside a "Business Economics" or policy group; OPERATIONS or INFORMATION SYSTEMS inside a "Technology & Operations" / "Decision Sciences" unit; managerial ACCOUNTING inside an "Accounting & X" unit.
- If such faculty EXIST: report present=true and give the department they sit in — reuse the EXACT name from the provided department list if it is one of those, otherwise give the school's own name for a newly-identified unit. (If the area is Technology and Operations, also set to_kind to "OM", "IS", or "OM+IS".)
- If the school GENUINELY has no faculty in that field: report present=false.

Output ONLY a JSON object:
{{"school":"<school>","results":[{{"ross_area":"<an unmapped area>","present":true,"dept":"<dept name>","to_kind":"OM|IS|OM+IS|","url":"<url>","confidence":"high|medium|low","notes":"<short; the journals>"}}, ...]}}
Include one result object for EVERY unmapped area you were given."""

# Shared, hardened roster procedure — the official live directory is the single source of truth
ROSTER_RULES = f"""SOURCE OF TRUTH = THE DEPARTMENT'S OWN OFFICIAL, CURRENT ONLINE FACULTY DIRECTORY. Accuracy and currency are non-negotiable; a roster padded with people who have left is a FAILURE. Follow this procedure EXACTLY:

1. LOCATE the official faculty / people / directory page for THIS SPECIFIC department on the school's OWN website (an official .edu / school domain). If a likely URL is provided, start there. Otherwise use the web_search tool ("<school> <department> faculty" or "<school> <department> faculty directory") and pick the school's OWN current listing — NEVER a third-party site, ranking, news article, Wikipedia, LinkedIn, ResearchGate, a faculty member's personal page, or an old/cached copy.

2. OPEN that directory page IN FULL with the fetch_page tool (pass the exact URL; do not rely on search snippets or your own memory for the roster). The people CURRENTLY listed there as ladder faculty ARE the roster.

3. EXHAUST EVERY PAGE — this is the #1 cause of missing faculty. Directories almost always show only a SLICE per page and split the rest across pagination, A–Z letters, rank tabs, or "load more" / "view all" controls. You MUST capture every slice:
   - URL-PARAMETER PAGINATION: if the page uses a query parameter such as page / p / pg / start / offset / from, iterate it (page=0, page=1, page=2, …) and call fetch_page on EACH page URL, following the directory's own "Next page" / numbered pagination links, until there are no more pages. The pagination control usually states the total number of pages — reach the LAST one. Do not stop at page 1.
   - Example pattern (a common Drupal directory): fetch_page on ".../directory?department=39&status=Active&page=0", then "&page=1", … — one fetch_page call per page until the pagination runs out.
   - Also iterate any A–Z last-name tabs or rank filters the directory uses.
   Keep calling fetch_page until you have enumerated the COMPLETE list across all pages/tabs.

4. PREFER AN ACTIVE / CURRENT FILTER when the directory offers one. If the directory has a status control (e.g. status=Active vs status=Emeritus, "current faculty", "active only"), use the ACTIVE/CURRENT value — it removes retired/emeritus/departed people at the source. IMPORTANT: an "Active" filter still typically INCLUDES non-ladder titles (lecturers, adjunct, clinical/teaching, professors of practice), so you must STILL apply the ladder-rank title filter in step 8 on top of it.

5. AUTHORITY: the current official directory governs both MEMBERSHIP and RANK. Read each person's rank from the title shown on that page.

6. CURRENCY (this is where rosters go wrong — get it right): include a person ONLY if they appear on the CURRENT official directory as active ladder faculty. Do NOT add anyone merely because your training memory, an older page, a CV, a paper, a Google Scholar profile, a news item, or a search snippet associates them with this department — those are routinely out of date. Anyone who has MOVED to another institution, RETIRED, or become EMERITUS must be EXCLUDED. If you are unsure about a name, verify it against the live directory and DROP it if it is not currently listed there.

7. COMPLETENESS: after capturing everyone across all pages, run a dedicated second pass for RECENTLY-HIRED / INCOMING ASSISTANT professors — the single most-missed group. Re-read the directory's full list, and search "<school> <department> assistant professor" and "<school> <department> new faculty {CURRENT_YEAR}" / "{CURRENT_YEAR - 1}". Add any CURRENT ladder faculty you find on the official directory that you missed. NEVER invent or guess a name.

8. INCLUDE ONLY tenured / tenure-track LADDER faculty, rank normalized to EXACTLY one of:
   - "Assistant Professor"
   - "Associate Professor"
   - "Professor"  (full professors, incl. named / endowed / chaired and University / Distinguished / Presidential professors)
   EXCLUDE everyone else and do NOT list them, EVEN IF they appear on the Active directory: lecturers, senior lecturers, professors of (the) practice, clinical / teaching professors, instructors, adjunct, visiting, affiliated / courtesy / secondary appointments from other units, postdocs, PhD students, and emeritus / retired / former faculty. (Read the title literally: "Lecturer of X" or "Adjunct Professor of X" or "Clinical Professor of X" → EXCLUDE; "Professor / Associate Professor / Assistant Professor of X" → include.)

9. TOOLS: use the web_search tool only to (a) find the official directory URL or (b) confirm an individual's CURRENT rank/affiliation; use fetch_page to OPEN the directory and every paginated page. web_search is NOT the roster source. Always prefer the school's own current pages; distrust undated or old-looking sources.

10. HONESTY: if you cannot retrieve the official directory in full (e.g. it will not load, or is JavaScript-only), set "complete" to false and explain briefly in "notes". Do NOT fill the gap from memory. Put the EXACT official directory URL(s) you used in "sources"."""

# Stage 2 — variant A: roster ONLY (unique-match department; area stamped by code)
FACULTY_SYSTEM_SIMPLE = f"""You are a meticulous research assistant compiling the COMPLETE, CURRENT LADDER-RANK faculty roster of ONE specific academic unit at ONE business school, strictly from that unit's own official directory.

{ROSTER_RULES}

Output ONLY a JSON object — no markdown:
{{"school":"<school>","area":"<unit>","complete":true,"sources":["<official directory url>", ...],"notes":"<short>","faculty":[{{"name":"<full name>","rank":"Assistant Professor|Associate Professor|Professor","title":"<raw title from the directory>"}}, ...]}}
Sort by rank (Professor, then Associate Professor, then Assistant Professor), then last name."""

# Stage 2 — variant B: roster + per-person Ross classification (flagged department)
FACULTY_SYSTEM_CLASSIFY = f"""You are a meticulous research assistant who (1) compiles the COMPLETE, CURRENT LADDER-RANK roster of ONE academic unit at ONE business school strictly from its official directory, and (2) classifies EACH faculty member into exactly ONE Ross area by the JOURNALS THAT PERSON publishes in.

This unit was flagged for individual review (its area-level mapping was not unique, or a human asked to verify it), so judge every person on their own record.

{ROSTER_RULES}

{FIELD_GUIDE}

PER-PERSON CLASSIFICATION RULES (apply only to the CURRENT ladder faculty you confirmed above):
- "ross_area" MUST be exactly one of: {", ".join(ROSS_AREAS)}.
- Decide it from the person's OWN research / the journals they publish in (check their faculty profile / Google Scholar via search), NEVER from the unit's name.
- The unit's likely Ross area(s) are given as a HINT, not a constraint — assign whatever truly fits each person, even outside the hint.
- "subfield": ONLY when ross_area is "Technology and Operations", set this to "OM" or "IS" by the person's journals — OM = Operations Management (Management Science(ops) / M&SOM / Production & Operations Management / Operations Research / Journal of Operations Management); IS = Information Systems (Information Systems Research / MIS Quarterly / Journal of MIS). For EVERY other Ross area, set "subfield" to "". OM and IS people publish in very different journals, so classify each individually.
- "field_evidence": <=14 words naming the journals/subfield that decided it (e.g. "ISR, MIS Quarterly -> IS").

Output ONLY a JSON object — no markdown:
{{"school":"<school>","area":"<unit>","complete":true,"sources":["<official directory url>", ...],"notes":"<short>","faculty":[{{"name":"<full name>","rank":"Assistant Professor|Associate Professor|Professor","title":"<raw title>","ross_area":"<one of the 7>","subfield":"OM|IS|","field_evidence":"<short>"}}, ...]}}
Sort by ross_area (in the 7-area order above), then rank, then last name."""

# Stage 2 — variant C: classify an ALREADY-CONFIRMED roster (two-step flow; does not re-enumerate)
FACULTY_CLASSIFY_LIST = f"""You are given the CONFIRMED CURRENT ladder-rank roster of ONE academic unit at ONE business school — it was already verified against the official directory. DO NOT change the roster: classify EXACTLY these people. Do not add, remove, merge, split, or rename anyone, and return every person you were given.

{FIELD_GUIDE}

For EACH listed person, assign their Ross area from THEIR OWN research and the JOURNALS THEY publish in (use web search to check their faculty profile / Google Scholar when unsure; spend your limited searches on the people you are least certain about — for obvious cases use what you already know).
- "ross_area" MUST be exactly one of: {", ".join(ROSS_AREAS)}. Judge by the person's field/journals, NOT the unit's name.
- "subfield": ONLY when ross_area is "Technology and Operations", set "OM" or "IS" by their journals (OM = Operations Management: Management Science(ops) / M&SOM / Production & Operations Management / Operations Research / Journal of Operations Management; IS = Information Systems: Information Systems Research / MIS Quarterly / Journal of MIS). For every other area, "".
- "field_evidence": <=14 words naming the journals/subfield that decided it.

Output ONLY a JSON object containing the SAME people, by their exact given names:
{{"school":"<school>","area":"<unit>","faculty":[{{"name":"<exact name as given>","rank":"<as given>","title":"<as given>","ross_area":"<one of the 7>","subfield":"OM|IS|","field_evidence":"<short>"}}, ...]}}"""

# ----------------------------------------------------------------------------- #
#  Cost meter                                                                    #
# ----------------------------------------------------------------------------- #
class CostMeter:
    def __init__(self, model):
        self.model, self.in_tok, self.out_tok, self.searches, self.calls = model, 0, 0, 0, 0
        self.lock = threading.Lock()

    def add(self, in_tok, out_tok, searches):
        with self.lock:
            self.in_tok += in_tok; self.out_tok += out_tok
            self.searches += searches; self.calls += 1

    def dollars(self):
        p = PRICE.get(self.model, {"in": 5.0, "out": 25.0})
        return (self.in_tok / 1e6) * p["in"] + (self.out_tok / 1e6) * p["out"] + (self.searches / 1000.0) * WEB_SEARCH_PRICE_PER_1K

    def summary(self):
        return (f"calls={self.calls}  in={self.in_tok:,}tok  out={self.out_tok:,}tok  "
                f"searches={self.searches}  ~= ${self.dollars():.2f}")

# ----------------------------------------------------------------------------- #
#  OpenAI call + JSON helpers                                                 #
# ----------------------------------------------------------------------------- #
def extract_json(text):
    """Return the LAST top-level JSON object found in the text. Robust to the narration the
    model emits between tool calls (which can contain stray braces and would break a naive
    first-'{'-to-last-'}' grab)."""
    if not text:
        return None
    t = re.sub(r"```(?:json)?", "", text, flags=re.I).strip()
    dec = json.JSONDecoder()
    best, i, n = None, 0, len(t)
    while i < n:
        if t[i] == "{":
            try:
                obj, end = dec.raw_decode(t[i:])
                if isinstance(obj, dict):
                    best = obj          # keep overwriting → ends as the final answer object
                    i += end
                    continue
            except ValueError:
                pass
        i += 1
    return best

def _usage_add(meter, resp):
    u = getattr(resp, "usage", None)
    in_tok = getattr(u, "input_tokens", 0) or 0
    out_tok = getattr(u, "output_tokens", 0) or 0
    searches = sum(1 for it in (getattr(resp, "output", None) or [])
                   if getattr(it, "type", "") == "web_search_call")
    meter.add(in_tok, out_tok, searches)

def _response_text(resp):
    t = getattr(resp, "output_text", "") or ""
    if t:
        return t
    parts = []
    for it in (getattr(resp, "output", None) or []):
        for c in (getattr(it, "content", None) or []):
            x = getattr(c, "text", None)
            if x:
                parts.append(x)
    return "\n".join(parts)

def call_model_json(client, meter, *, system, user, model, use_web_fetch, max_tokens, tries=3, max_tool_turns=8):
    """OpenAI Responses API call with an agentic loop:
       - the hosted web_search tool runs automatically inside each response (discovery/confirmation);
       - our fetch_page() function calls are executed here and fed back (deterministic page fetching
         + pagination), chained via previous_response_id;
       - output is forced to a single JSON object to avoid the prose/JSON-mix parse failures.
    Retries only TRANSIENT (pre-billing) errors; a billed-but-unparseable response is NOT retried."""
    tools = [WEB_SEARCH_TOOL] + ([FETCH_TOOL] if use_web_fetch else [])
    base = dict(model=model, instructions=system, tools=tools,
                max_output_tokens=max_tokens, text={"format": {"type": "json_object"}})
    last = None
    for attempt in range(tries):
        try:
            resp = client.responses.create(input=user, **base)
        except Exception as e:
            last = e  # transient (network / 429 / 5xx) — safe and cheap to retry from scratch
            if attempt < tries - 1:
                time.sleep(2 ** attempt)
                continue
            raise
        # --- agentic tool loop: satisfy fetch_page calls until the model returns its answer ---
        turns = 0
        while True:
            _usage_add(meter, resp)
            calls = [it for it in (getattr(resp, "output", None) or [])
                     if getattr(it, "type", "") == "function_call" and getattr(it, "name", "") == "fetch_page"]
            if not calls:
                break
            if turns >= max_tool_turns:  # stop fetching; force a final JSON answer with no tools
                stop = [{"type": "function_call_output", "call_id": c.call_id, "output": "(fetch budget exhausted)"} for c in calls]
                cfg = {k: v for k, v in base.items() if k != "tools"}
                try:
                    resp = client.responses.create(input=stop, previous_response_id=resp.id, tools=[], **cfg)
                    _usage_add(meter, resp)
                except Exception:
                    pass
                break
            outputs = []
            for c in calls:
                try:
                    args = json.loads(c.arguments or "{}")
                except Exception:
                    args = {}
                outputs.append({"type": "function_call_output", "call_id": c.call_id,
                                "output": fetch_page(args.get("url", ""))})
            turns += 1
            resp = client.responses.create(input=outputs, previous_response_id=resp.id, **base)
        # --- parse the final JSON ---
        parsed = extract_json(_response_text(resp))
        if parsed is not None:
            return parsed
        if getattr(resp, "status", "") == "incomplete":
            raise ValueError("the response was cut off before its JSON finished (hit the token limit) — "
                             "this department's roster may be unusually large; re-run just this one.")
        raise ValueError("the AI's response wasn't valid JSON. Not retried automatically (to avoid "
                         "re-paying for the web work) — re-run this department.")
    raise last

# ----------------------------------------------------------------------------- #
#  Ross-area normalization + department helpers                                  #
# ----------------------------------------------------------------------------- #
def is_no_match(t):
    return (not t) or bool(re.search(r"no direct|no ross|^none$|^n/?a$|no equivalent", str(t).strip(), re.I))

# Ordered so Strategy / T&O / BEPP are matched before the broad "management".
_AREA_RULES = [
    (r"account", "Accounting"),
    (r"financ", "Finance"),
    (r"market", "Marketing"),
    (r"strateg", "Strategy"),
    (r"operation|supply chain|manufactur|information system|info\.?\s*system|decision scien|\bmis\b|"
     r"information technolog|technolog", "Technology and Operations"),
    (r"econ|public polic|managerial economic|\bbepp\b|law and economic|law & economic", "Business Economics and Public Policy"),
    (r"manage|organization|organis|behav|human resource|\bhr\b|\bob\b|people analytics", "Management and Organizations"),
]
def normalize_area(s):
    if not s or is_no_match(s):
        return None
    if s in ROSS_AREAS:
        return s
    low = str(s).lower()
    for pat, area in _AREA_RULES:
        if re.search(pat, low):
            return area
    return None

def valid_ross(areas):
    """Normalize a list (or ';'-joined string) of Ross-area labels to canonical names, deduped."""
    if isinstance(areas, str):
        areas = re.split(r"\s*;\s*", areas)
    out = []
    for a in (areas or []):
        n = normalize_area(a)
        if n and n not in out:
            out.append(n)
    return out

def normalize_to_kind(s):
    """Normalize a department-level T&O subfield tag to 'OM' | 'IS' | 'OM+IS' | ''."""
    t = str(s or "").strip().lower()
    has_om = bool(re.search(r"\bom\b|operation|supply chain|manufactur", t))
    has_is = bool(re.search(r"\bis\b|information system|info\.?\s*system|\bmis\b|information technolog|decision scien", t))
    if "om+is" in t.replace(" ", "") or (has_om and has_is):
        return "OM+IS"
    if has_om:
        return "OM"
    if has_is:
        return "IS"
    return ""

def normalize_subfield(s):
    """Per-person T&O subfield: 'OM' | 'IS' | '' (never both for one person)."""
    t = str(s or "").strip().lower()
    if re.search(r"\bis\b|information system|info\.?\s*system|\bmis\b|information technolog", t):
        return "IS"
    if re.search(r"\bom\b|operation|supply chain|manufactur", t):
        return "OM"
    return ""

def default_force_check(ross_areas, to_kind=""):
    """Per-faculty checking is ON by default when the department is NOT a unique match,
    OR when it is a combined OM+IS unit (so OM vs IS can be split per person)."""
    return len(ross_areas) >= 2 or normalize_to_kind(to_kind) == "OM+IS"

def parse_departments(parsed):
    """Normalize a Stage-1 model response into a clean department list."""
    deps = []
    for d in (parsed.get("departments") or []):
        name = (d.get("dept") or d.get("name") or "").strip()
        if not name:
            continue
        ross = valid_ross(d.get("ross_areas", []))
        to_kind = normalize_to_kind(d.get("to_kind", "")) if ("Technology and Operations" in ross) else ""
        deps.append({"name": name, "ross_areas": ross, "to_kind": to_kind,
                     "url": (d.get("url") or "").strip(),
                     "confidence": (d.get("confidence") or "").strip().lower(),
                     "notes": (d.get("notes") or "").strip(),
                     "force_check": default_force_check(ross, to_kind)})
    return deps

def ross_coverage(departments):
    """Which of the 7 Ross areas have at least one mapped department. Returns {area: bool}."""
    covered = set()
    for d in departments:
        covered.update(d["ross_areas"])
    return {a: (a in covered) for a in ROSS_AREAS}

def merge_coverage(deps, cov_parsed, gaps):
    """Fold an automatic coverage-recovery result into the department list.
    Returns (deps, confirmed_absent_areas)."""
    by = {d["name"].lower(): d for d in deps}
    absent = []
    for r in (cov_parsed.get("results") or []):
        area = normalize_area(r.get("ross_area"))
        if not area or area not in gaps:
            continue
        if bool(r.get("present")):
            name = (r.get("dept") or "").strip()
            tk = normalize_to_kind(r.get("to_kind", "")) if area == "Technology and Operations" else ""
            if not name:
                continue
            ex = by.get(name.lower())
            if ex:
                if area not in ex["ross_areas"]:
                    ex["ross_areas"].append(area)
                if area == "Technology and Operations" and tk and not ex.get("to_kind"):
                    ex["to_kind"] = tk
            else:
                nd = {"name": name, "ross_areas": [area], "to_kind": tk,
                      "url": (r.get("url") or "").strip(), "confidence": (r.get("confidence") or "").strip().lower(),
                      "notes": (r.get("notes") or "").strip(), "force_check": False}
                deps.append(nd); by[name.lower()] = nd
        else:
            absent.append(area)
    for d in deps:  # recompute after possible area growth
        d["force_check"] = default_force_check(d["ross_areas"], d.get("to_kind", ""))
    return deps, absent

LADDER = {"Professor", "Associate Professor", "Assistant Professor"}

def _name_key(n):
    """Normalize a person's name for matching across the roster and classification steps."""
    return re.sub(r"[^a-z ]", "", str(n or "").lower()).strip()

def fetch_department_faculty(client, meter, model, use_fetch, s, d, *, recheck=False, prior=None,
                             max_tokens_roster=10000, max_tokens_classify=12000):
    """Two-step faculty fetch.
    STEP 1 (always): build the COMPLETE CURRENT roster from the official directory (same hardened
    procedure for every department). STEP 2 (only when the department is flagged): classify each
    CONFIRMED person by their journals, without changing the roster. Unique-match departments skip
    step 2 and inherit their single Ross area (and OM/IS subfield if known).
    Returns {complete, sources, notes, faculty:[{name,rank,title,ross_area,subfield,field_evidence}], classified}.
    """
    classify = bool(d["force_check"])
    stamp = d["ross_areas"][0] if len(d["ross_areas"]) == 1 else ""
    stamp_sub = d.get("to_kind", "") if (stamp == "Technology and Operations" and d.get("to_kind") in ("OM", "IS")) else ""
    hint = f"Official directory page to start from (fetch this first): {d['url']}\n" if d.get("url") else ""
    prior_block = ""
    if recheck and prior:
        prior_block = ("A PRIOR attempt produced this roster, possibly INCOMPLETE or out of date:\n"
                       + "\n".join(f"- {p.get('name')} ({p.get('rank')})" for p in prior)
                       + "\nRe-open the OFFICIAL directory: ADD any current ladder faculty missing (especially new "
                         "ASSISTANT professors); REMOVE anyone not currently listed there (moved away, retired, emeritus, non-ladder).\n")

    # ---- STEP 1: roster from the official directory ----
    user1 = (f"School: {s['university']} — {s['school']}\nUnit (the school's own name): {d['name']}\n{hint}{prior_block}"
             "Use fetch_page to open the unit's OFFICIAL faculty directory (walking every page of pagination) and build "
             "the roster ONLY from who is CURRENTLY listed there (exclude anyone who has moved away, retired, or is "
             "emeritus). Be exhaustive — do not miss new assistant professors. Return ONLY the JSON.")
    parsed = call_model_json(client, meter, system=FACULTY_SYSTEM_SIMPLE, user=user1, model=model,
                              use_web_fetch=use_fetch, max_tokens=max_tokens_roster)
    roster = [{"name": f.get("name", ""), "rank": f.get("rank", ""), "title": f.get("title", "")}
              for f in parsed.get("faculty", []) if f.get("rank") in LADDER]

    if not classify:
        faculty = [{**p, "ross_area": stamp, "subfield": stamp_sub, "field_evidence": ""} for p in roster]
    elif not roster:
        faculty = []
    else:
        # ---- STEP 2: classify exactly the confirmed roster ----
        names_block = "\n".join(f"- {p['name']} ({p['rank']})" for p in roster)
        split = ("This is a combined Operations + Information Systems unit — be sure to tag each T&O person OM or IS.\n"
                 if d.get("to_kind") == "OM+IS" else "")
        cand = ", ".join(d["ross_areas"]) if d["ross_areas"] else "(unconstrained)"
        user2 = (f"School: {s['university']} — {s['school']}\nUnit: {d['name']}\n"
                 f"The unit's likely Ross area(s) (context only, judge each person independently): {cand}\n{split}"
                 f"CONFIRMED CURRENT roster to classify (return exactly these people, unchanged):\n{names_block}\n"
                 "Classify each person by the journals THEY publish in. Return ONLY the JSON.")
        cls = {}
        try:
            p2 = call_model_json(client, meter, system=FACULTY_CLASSIFY_LIST, user=user2, model=model,
                                  use_web_fetch=use_fetch, max_tokens=max_tokens_classify)
            for f in p2.get("faculty", []):
                area = normalize_area(f.get("ross_area"))
                sub = normalize_subfield(f.get("subfield")) if area == "Technology and Operations" else ""
                cls[_name_key(f.get("name"))] = (area, sub, f.get("field_evidence", ""))
        except Exception:
            cls = {}
        faculty = []
        for p in roster:
            area, sub, ev = cls.get(_name_key(p["name"]), (None, "", ""))
            if not area:
                area = stamp or "Unclassified"  # fall back to the unique area, else leave for review
                sub = stamp_sub if area == "Technology and Operations" else ""
            faculty.append({**p, "ross_area": area, "subfield": sub, "field_evidence": ev})

    return {"complete": bool(parsed.get("complete")), "sources": parsed.get("sources", []),
            "notes": parsed.get("notes", ""), "faculty": faculty, "classified": classify}

# ----------------------------------------------------------------------------- #
#  Resumable cache                                                               #
# ----------------------------------------------------------------------------- #
class Cache:
    def __init__(self, enabled, root=".ross_cache"):
        self.enabled, self.root = enabled, Path(root)
        if enabled:
            self.root.mkdir(parents=True, exist_ok=True)

    def _path(self, key):
        return self.root / (hashlib.sha1(key.encode("utf-8")).hexdigest()[:16] + ".json")

    def get(self, key):
        if not self.enabled:
            return None
        p = self._path(key)
        if p.exists():
            try:
                return json.loads(p.read_text())
            except Exception:
                return None
        return None

    def put(self, key, value):
        if self.enabled:
            try:
                self._path(key).write_text(json.dumps(value))
            except Exception:
                pass

# ----------------------------------------------------------------------------- #
#  Table I/O                                                                     #
# ----------------------------------------------------------------------------- #
def _read_table(path):
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Reading .xlsx needs openpyxl:  pip install openpyxl  (or export to .csv)")
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(c) if c is not None else "" for c in rows[0]]
        return [{header[i]: ("" if v is None else v) for i, v in enumerate(r)} for r in rows[1:]]
    with open(p, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))

def _pick(keys, *patterns):
    for pat in patterns:
        for k in keys:
            if re.search(pat, str(k), re.I):
                return k
    return None

def read_school_list(path):
    if not path:
        return [{"ranking": r, "university": u, "school": s} for r, u, s in BUILTIN]
    rows = _read_table(path)
    out = []
    for row in rows:
        keys = list(row.keys())
        rk = _pick(keys, r"rank"); uk = _pick(keys, r"univ")
        sk = _pick(keys, r"business\s*school", r"^school$", r"b-?school", r"school")
        s = {"ranking": str(row.get(rk, "")).strip() if rk else "",
             "university": str(row.get(uk, "")).strip() if uk else "",
             "school": str(row.get(sk, "")).strip() if sk else ""}
        if s["university"] or s["school"]:
            out.append(s)
    if not out:
        sys.exit(f"No school rows found in {path}.")
    return out

def schools_from_stage1_rows(rows):
    """Build per-school department lists from Stage-1 CSV rows.
    Supports the v4 department-centric format AND the old Ross-area-centric format."""
    if not rows:
        return []
    keys = list(rows[0].keys())
    dept_col = _pick(keys, r"school\s*department", r"department.*own", r"^department$", r"unit")
    if dept_col:  # ---- v4 format ----
        rk = _pick(keys, r"rank"); uk = _pick(keys, r"univ")
        sk = _pick(keys, r"business\s*school", r"^school$", r"b-?school", r"school$")
        rossk = _pick(keys, r"ross\s*area"); ck = _pick(keys, r"check\s*fac|individual")
        urlk = _pick(keys, r"url|source"); confk = _pick(keys, r"conf")
        tkk = _pick(keys, r"t&o\s*subfield", r"t.o\s*subfield", r"subfield")
        by = {}
        for row in rows:
            uni = str(row.get(uk, "")).strip() if uk else ""
            sch = str(row.get(sk, "")).strip() if sk else ""
            dept = str(row.get(dept_col, "")).strip()
            if not dept or not (uni or sch):
                continue
            entry = by.setdefault((uni, sch), {"ranking": str(row.get(rk, "")).strip() if rk else "",
                                               "university": uni, "school": sch, "departments": []})
            ross = valid_ross(str(row.get(rossk, "")) if rossk else "")
            to_kind = normalize_to_kind(str(row.get(tkk, ""))) if (tkk and "Technology and Operations" in ross) else ""
            force = None
            if ck:
                force = str(row.get(ck, "")).strip().lower() in ("yes", "true", "1", "on")
            entry["departments"].append({
                "name": dept, "ross_areas": ross, "to_kind": to_kind,
                "url": (str(row.get(urlk, "")).strip() if urlk else ""),
                "confidence": (str(row.get(confk, "")).strip().lower() if confk else ""),
                "notes": "", "force_check": force if force is not None else default_force_check(ross, to_kind)})
        return list(by.values())
    # ---- old Ross-area-centric format: invert ----
    rk = _pick(keys, r"rank"); uk = _pick(keys, r"univ")
    sk = _pick(keys, r"business\s*school", r"^school$", r"b-?school", r"school$")
    matchk = _pick(keys, r"match"); urlk = _pick(keys, r"url|source")
    rossk = _pick(keys, r"ross")
    by = {}
    for row in rows:
        uni = str(row.get(uk, "")).strip() if uk else ""
        sch = str(row.get(sk, "")).strip() if sk else ""
        if not (uni or sch):
            continue
        entry = by.setdefault((uni, sch), {"ranking": str(row.get(rk, "")).strip() if rk else "",
                                           "university": uni, "school": sch, "_deps": {}})
        matched = str(row.get(matchk, "")).strip() if matchk else ""
        ross_area = normalize_area(str(row.get(rossk, ""))) if rossk else None
        if is_no_match(matched) or not ross_area:
            continue
        for part in re.split(r"\s*;\s*", matched):
            nm = part.strip()
            if not nm or is_no_match(nm):
                continue
            dk = nm.lower()
            d = entry["_deps"].setdefault(dk, {"name": nm, "ross_areas": [], "url": "", "confidence": "", "notes": ""})
            if ross_area not in d["ross_areas"]:
                d["ross_areas"].append(ross_area)
            if not d["url"] and urlk and row.get(urlk):
                d["url"] = str(row.get(urlk)).strip()
    out = []
    for e in by.values():
        deps = [{**d, "to_kind": "", "force_check": default_force_check(d["ross_areas"])} for d in e["_deps"].values()]
        out.append({"ranking": e["ranking"], "university": e["university"], "school": e["school"], "departments": deps})
    return out

def csv_cell(v):
    return "" if v is None else v

# ----------------------------------------------------------------------------- #
#  Stage 1 — area matching (reversed)                                            #
# ----------------------------------------------------------------------------- #
def stage1(args):
    client = OpenAI()
    meter = CostMeter(args.model)
    cache = Cache(not args.no_cache, root=".ross_cache/match")
    schools = read_school_list(args.input)
    if args.limit:
        schools = schools[: args.limit]
    print(f"Stage 1 — mapping {len(schools)} schools (each school's units -> Ross areas) "
          f"[model={args.model}, web_fetch={'on' if not args.no_web_fetch else 'off'}, conc={args.concurrency}]")
    results = {}

    def work(s):
        key = f"matchR2::{args.model}::{s['university']}||{s['school']}"
        cached = cache.get(key)
        if cached is not None:
            return s, cached["deps"], cached.get("absent", []), True
        user = (f"Target school: {s['university']} — {s['school']}.\n"
                "Enumerate this school's OWN academic units (departments / areas / faculty research groups) from its "
                "official site, and map EACH unit to the Ross area(s) it fits, judging by the journals its faculty publish in. "
                "A unit may map to one Ross area, several, or none. Apply the Operations/Information-Systems rule, set to_kind "
                "for T&O units, and run the mandatory per-area coverage check before returning. Return ONLY the JSON.")
        parsed = call_model_json(client, meter, system=MATCH_SYSTEM, user=user, model=args.model,
                                  use_web_fetch=not args.no_web_fetch, max_tokens=3000)
        deps = parse_departments(parsed)
        absent = []
        gaps = [a for a, ok in ross_coverage(deps).items() if not ok]
        if gaps:  # automatic immediate re-check for the missing areas
            dep_list = "; ".join(d["name"] for d in deps) or "(none found)"
            cu = (f"School: {s['university']} — {s['school']}\n"
                  f"Departments already found: {dep_list}\n"
                  f"Ross areas still UNMAPPED: {', '.join(gaps)}\n"
                  "For each unmapped area, find the department whose faculty cover it (often a combined unit, e.g. Strategy "
                  "inside a Management department) and report present=true with that department's name, or present=false if the "
                  "school truly lacks it. Judge by journals. Return ONLY the JSON.")
            try:
                cov = call_model_json(client, meter, system=COVERAGE_SYSTEM, user=cu, model=args.model,
                                       use_web_fetch=not args.no_web_fetch, max_tokens=1500)
                deps, absent = merge_coverage(deps, cov, gaps)
            except Exception:
                pass
        cache.put(key, {"deps": deps, "absent": absent})
        return s, deps, absent, False

    with cf.ThreadPoolExecutor(max_workers=args.concurrency) as ex:
        futs = {ex.submit(work, s): s for s in schools}
        done = 0
        for fut in cf.as_completed(futs):
            s = futs[fut]
            try:
                s, deps, absent, hit = fut.result()
                results[(s["university"], s["school"])] = (s, deps)
                miss = [ROSS_SHORT[a] for a, ok in ross_coverage(deps).items() if not ok]
                done += 1
                flag = f"  [absent: {','.join(miss)}]" if miss else ""
                print(f"  [{done}/{len(schools)}] {'cache' if hit else 'live '} {s['school'][:36]:36s} "
                      f"{len(deps)} units{flag}  ${meter.dollars():.2f}")
            except Exception as e:
                done += 1
                print(f"  [{done}/{len(schools)}] FAIL  {s['school'][:36]:36s}  {e}")
                results[(s["university"], s["school"])] = (s, None)

    _write_stage1_csv(args.out, schools, results)
    print(f"\nWrote {args.out}.  Total {meter.summary()}")
    return args.out

def _write_stage1_csv(path, schools, results):
    header = ["Ranking", "University", "Business School", "School Department", "Ross Area(s)",
              "Unique Match", "Check Faculty Individually", "T&O Subfield", "Confidence", "Source URL", "Notes"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for s in schools:
            _, deps = results.get((s["university"], s["school"]), (s, None))
            for d in (deps or []):
                ross = " ; ".join(d["ross_areas"]) if d["ross_areas"] else "No Ross equivalent"
                unique = "yes" if len(d["ross_areas"]) == 1 else "no"
                w.writerow([s["ranking"], s["university"], s["school"], d["name"], ross, unique,
                            "yes" if d["force_check"] else "no", d.get("to_kind", ""), d.get("confidence", ""),
                            d.get("url", ""), d.get("notes", "")])
            if not deps:
                w.writerow([s["ranking"], s["university"], s["school"], "", "", "", "", "", "", "", "lookup failed"])

# ----------------------------------------------------------------------------- #
#  Stage 2 — faculty (conditional per-person classification)                     #
# ----------------------------------------------------------------------------- #
def stage2(args):
    client = OpenAI()
    meter = CostMeter(args.model)
    cache = Cache(not args.no_cache, root=".ross_cache/faculty")
    schools = schools_from_stage1_rows(_read_table(args.infile))
    if args.limit:
        schools = schools[: args.limit]
    jobs = []
    for s in schools:
        for d in s["departments"]:
            if not d["ross_areas"] and not args.include_unmapped:
                continue  # skip "No Ross equivalent" units unless asked
            jobs.append((s, d))
    n_classify = sum(1 for _s, d in jobs if d["force_check"])
    print(f"Stage 2 — {len(schools)} schools, {len(jobs)} departments "
          f"({n_classify} check faculty individually, {len(jobs) - n_classify} inherit their area) "
          f"[model={args.model}, web_fetch={'on' if not args.no_web_fetch else 'off'}, conc={args.concurrency}]")
    results = {}

    def work(s, d):
        classify = bool(d["force_check"])
        key = f"facR3::{args.model}::{'C' if classify else 'S'}::{s['university']}||{s['school']}||{d['name']}"
        cached = cache.get(key)
        if cached is not None:
            return s, d, cached, True
        data = fetch_department_faculty(client, meter, args.model, not args.no_web_fetch, s, d)
        cache.put(key, data)
        return s, d, data, False

    with cf.ThreadPoolExecutor(max_workers=args.concurrency) as ex:
        futs = {ex.submit(work, s, d): (s, d) for s, d in jobs}
        done = 0
        for fut in cf.as_completed(futs):
            s, d = futs[fut]
            try:
                s, d, data, hit = fut.result()
                results[(s["university"], s["school"], d["name"])] = data
                done += 1
                tag = "cache" if hit else ("clf  " if data["classified"] else "stamp")
                print(f"  [{done}/{len(jobs)}] {tag} {'ok ' if data['complete'] else 'PART'} "
                      f"{len(data['faculty']):>3} fac  {(s['school']+' / '+d['name'])[:46]:46s}  ${meter.dollars():.2f}")
            except Exception as e:
                done += 1
                print(f"  [{done}/{len(jobs)}] FAIL  {(s['school']+' / '+d['name'])[:46]:46s}  {e}")
                results[(s["university"], s["school"], d["name"])] = None

    _write_stage2_csv(args.out, schools, results, args.include_unmapped)
    print(f"\nWrote {args.out}.  Total {meter.summary()}")

def _write_stage2_csv(path, schools, results, include_unmapped):
    header = ["Ranking", "University", "Business School", "School Department", "Dept Ross Area(s)",
              "Checked Individually", "Faculty Name", "Rank", "Faculty Ross Area", "T&O Subfield",
              "Field Evidence", "Complete", "Sources"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for s in schools:
            for d in s["departments"]:
                if not d["ross_areas"] and not include_unmapped:
                    continue
                data = results.get((s["university"], s["school"], d["name"]))
                dept_ross = " ; ".join(d["ross_areas"]) if d["ross_areas"] else "No Ross equivalent"
                clf = "yes" if d["force_check"] else "no"
                if not data:
                    w.writerow([s["ranking"], s["university"], s["school"], d["name"], dept_ross, clf,
                                "", "", "", "", "", "ERROR", ""]); continue
                sources = " ; ".join(data.get("sources", []) or [])
                comp = "yes" if data.get("complete") else "no"
                fac = data.get("faculty", [])
                if fac:
                    for p in fac:
                        w.writerow([s["ranking"], s["university"], s["school"], d["name"], dept_ross, clf,
                                    p.get("name", ""), p.get("rank", ""), p.get("ross_area", ""),
                                    p.get("subfield", ""), p.get("field_evidence", ""), comp, sources])
                else:
                    w.writerow([s["ranking"], s["university"], s["school"], d["name"], dept_ross, clf,
                                "", "", "", "", "", comp, sources])

# ----------------------------------------------------------------------------- #
#  CLI                                                                           #
# ----------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="Ross Area & Faculty Toolkit (v4, reversed mapping)")
    sub = ap.add_subparsers(dest="cmd", required=True)

    def common(p):
        p.add_argument("--model", default=DEFAULT_MODEL, help=f"default {DEFAULT_MODEL}; cheaper: {SONNET}")
        p.add_argument("--concurrency", type=int, default=3)
        p.add_argument("--no-web-fetch", action="store_true")
        p.add_argument("--no-cache", action="store_true")
        p.add_argument("--limit", type=int, default=0)

    m = sub.add_parser("match", help="Stage 1 — map each school's units to Ross areas")
    common(m)
    m.add_argument("--input", default="", help="schools .xlsx/.csv; omit for built-in Top 50")
    m.add_argument("--out", default="ross_area_matches.csv")

    fa = sub.add_parser("faculty", help="Stage 2 — find + (conditionally) classify faculty")
    common(fa)
    fa.add_argument("--in", dest="infile", default="ross_area_matches.csv")
    fa.add_argument("--out", default="ross_faculty_by_area.csv")
    fa.add_argument("--include-unmapped", action="store_true", help="also pull faculty for 'No Ross equivalent' units")

    al = sub.add_parser("all", help="run Stage 1 then Stage 2")
    common(al)
    al.add_argument("--input", default="")
    al.add_argument("--include-unmapped", action="store_true")

    args = ap.parse_args()
    if OpenAI is None:
        sys.exit("Missing dependency. Run:  pip install -r requirements.txt")
    if not os.environ.get("OPENAI_API_KEY"):
        sys.exit("Set your key first:  export OPENAI_API_KEY=sk-...")

    if args.cmd == "match":
        stage1(args)
    elif args.cmd == "faculty":
        stage2(args)
    elif args.cmd == "all":
        args.out = "ross_area_matches.csv"; stage1(args)
        args.infile = "ross_area_matches.csv"; args.out = "ross_faculty_by_area.csv"; stage2(args)

if __name__ == "__main__":
    main()
